# coding=utf-8
from __future__ import unicode_literals
from django.contrib import messages
from django.db import DatabaseError
from django.http import HttpResponse
from django.shortcuts import redirect
from django.views.decorators.cache import never_cache
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from wkhtmltopdf.views import PDFTemplateView
import xlwt
from django_apogee.models import InsAdmEtp, Individu, InsAdmEtpInitial
from xadmin.views import filter_hook, BaseAdminView
from xadmin import views
import xadmin
import datetime
from duck_inscription.models import SettingsEtape, NoteMasterModel
from duck_inscription.models import Wish

try:
    from foad.utils import make_etudiant_password
except ImportError:
    pass


class StatistiquePal(views.Dashboard):
    base_template = 'statistique/stats_pal.html'
    widget_customiz = False

    def get_context(self):
        context = super(StatistiquePal, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_pal'), 'title': 'Statistique PAL'}]

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())
xadmin.site.register_view(r'^stats_pal/$', StatistiquePal, 'stats_pal')


class StatistiquePiel(views.Dashboard):
    base_template = 'statistique/stats_piel.html'
    widget_customiz = False

    @filter_hook
    def get_context(self):
        context = super(StatistiquePiel, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_piel'), 'title': 'Statistique PIEL'}]
xadmin.site.register_view(r'^stats_piel/$', StatistiquePiel, 'stats_piel')


class StatistiqueApogee(views.Dashboard):
    base_template = 'statistique/stats_apogee.html'
    widget_customiz = False

    @filter_hook
    def get_context(self):
        context = super(StatistiqueApogee, self).get_context()
        context['years'] = [str(x) for x in range(2010, datetime.date.today().year+1)]
        context['selected'] = str(self.request.GET.get('year', datetime.date.today().year))
        etapes = []
        total = 0
        for etape in SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome'):
            query = InsAdmEtpInitial.objects.using('oracle').filter(cod_anu=context['selected'], eta_iae='E',
                                                                              cod_pru__in=['NO', 'FP', 'DD'], tem_iae_prm='O',
                                                                              cod_etp=etape.cod_etp) | InsAdmEtpInitial.objects.using('oracle').filter(
                cod_anu=context['selected'], eta_iae='E', tem_iae_prm='O', cod_etp=etape.cod_etp)

            etapes.append({'label': etape.label,
                           'cod_etp': etape.cod_etp,
                           'nb_inscrit': query.count()})
            total += query.count()

        context['etapes'] = etapes
        context['total'] = total
        return context

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_apogee'), 'title': 'Statistique Apogee'}]
xadmin.site.register_view(r'^stats_apogee/$', StatistiqueApogee, 'stats_apogee')


class ExtractionStatistiqueBase(BaseAdminView):

    def set_attr_queryset(self, **kwargs):
        for attr in self.attrs_queryset:
            setattr(self, attr, kwargs.get(attr, None))

    @property
    def get_structure_excel(self):
        return self.structure_excel

    def get_queryset(self):
        return self.model_extraction.objects.filter(**self.filter_queryset[self.type_stat])

    def create_workbook(self):
        wb = Workbook()
        ws = wb.active

        queryset = self.get_queryset()
        for collumn, cell in enumerate(self.get_structure_excel, start=1):
            ws.cell(row=1, column=collumn).value = cell[0]

        for row, obj in enumerate(queryset, start=2):
            for collumn, cell in enumerate(self.get_structure_excel, start=1):
                try:
                    ws.cell(row=row, column=collumn).value = eval(cell[1])
                except NameError:
                    pass
        return wb

    def get(self, request, *args, **kwargs):
        self.set_attr_queryset(**kwargs)
        response = HttpResponse(save_virtual_workbook(self.create_workbook()), content_type='application/vnd.ms-excel')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xlsx' % ('extraction', date)
        return response


class ExtrationStatistique(ExtractionStatistiqueBase):
    model_extraction = Wish
    attrs_queryset = ['step', 'etat', 'type_stat']
    structure_excel = [['Numero Etudiant', 'obj.individu.student_code'],
                       ['Nom patronimique', 'obj.individu.last_name'],
                       ['Nom d\'époux', 'obj.individu.common_name'],
                       ["Prénom:", 'obj.individu.first_name1'],
                       ['Deuxiéme prénom', "obj.individu.first_name2"],
                       ["Email", 'obj.individu.personal_email']]

    @property
    def filter_queryset(self):
        return {
            'parcours_dossier': {'etape__cod_etp': self.step, 'parcours_dossier__to_state': self.etat},
            'state': {'state': self.etat, ' etape__cod_etp': self.step},
            'etat': {'suivi_dossier': self.etat, 'etape__cod_etp': self.step},
            'etat_suivi_dossier': {'etape_dossier__to_state': self.etat, 'etape__cod_etp': self.step}
        }
xadmin.site.register_view(r'^extraction/(?P<type_stat>\w+)/(?P<etat>\w+)/(?P<step>\w+)/$', ExtrationStatistique,
                          'extraction_stat')


class ExtractionStatApogee(ExtractionStatistiqueBase):

    model_extraction = InsAdmEtpInitial
    attrs_queryset = ['step', 'annee']
    structure_excel = [['Numero Etudiant', 'obj.cod_ind.cod_etu'],
                       ['Nom patronimique', 'obj.cod_ind.lib_nom_pat_ind'],
                       ['Nom d\'époux', 'obj.cod_ind.lib_nom_usu_ind'],
                       ["Prénom:", 'obj.cod_ind.lib_pr1_ind'],
                       ['Deuxiéme prénom', "obj.cod_ind.lib_pr2_ind"],
                       ["Email perso", 'str(obj.cod_ind.get_email(self.annee))'],
                       ["Email Foad", 'str(obj.cod_ind.cod_etu) + \'@foad.iedparis8.net\''],
                       ["Reinscription:", "'Oui' if obj.is_reins else 'Non'"],
                       ["Password:", "str(make_etudiant_password(obj.cod_ind.cod_etu))"],
                       ]

    type_stat = 'ordinaire'
    url = '/stats_apogee/'

    def get_queryset(self):
        return self.model_extraction.inscrits.using('oracle').filter(**self.filter_queryset[self.type_stat])

    def get(self, request, *args, **kwargs):
        self.set_attr_queryset(**kwargs)
        try:
            response = HttpResponse(save_virtual_workbook(self.create_workbook()), content_type='application/vnd.ms-excel')
        except DatabaseError:
            messages.error(request, 'Connection à apogée impossible')
            return redirect('xadmin:stat_apogee')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xlsx' % ('extraction', date)
        return response

    @property
    def filter_queryset(self):
        return {
            'ordinaire': {'cod_anu': self.annee, 'cod_etp': self.step},
        }
xadmin.site.register_view(r'^extraction_apogee/(?P<annee>\w+)/(?P<step>\w+)/$', ExtractionStatApogee,
                          'extraction_apogee')


class ExtractionPiel(views.Dashboard):
    base_template = 'extraction/extraction_pal.html'
    widget_customiz = False

    @filter_hook
    def get_context(self):
        context = super(ExtractionPiel, self).get_context()
        context['etapes'] = SettingsEtape.objects.filter(is_inscription_ouverte=True).order_by('diplome')
        return context

    @never_cache
    def get(self, request, *args, **kwargs):
        self.widgets = self.get_widgets()
        return self.template_response(self.base_template, self.get_context())

    @filter_hook
    def get_breadcrumb(self):
        return [{'url': self.get_admin_url('index'), 'title': 'Accueil'},
                {'url': self.get_admin_url('statistiques'), 'title': 'Statistique'},
                {'url': self.get_admin_url('stats_piel'), 'title': 'Statistique PIEL'}]


xadmin.site.register_view(r'^extraction/$', ExtractionPiel, 'extraction')


class DatesAndTarifs(ExtractionPiel):
    base_template = "informations/date_et_tarifs.html"

xadmin.site.register_view(r'^datesandtarifs/$', DatesAndTarifs, 'datesandtarifs')


class ExtrationEtiquetteView(PDFTemplateView):
    filename = "etiquette.pdf"
    template_name = "extraction/etiquette_postale.html"
    cmd_options = {
        'orientation': 'landscape',
    }

    def get_context_data(self, **kwargs):
        cod_ind = self.kwargs.get('cod_ind', None)
        etudiant = Individu.objects.get(cod_ind=cod_ind)
        context = super(ExtrationEtiquetteView, self).get_context_data(**kwargs)
        context['etudiants'] = [etudiant] * 4
        return context


class ExtrationEtiquettesView(PDFTemplateView):
    filename = "Etiquettes.pdf"
    template_name = "extraction/etiquette_postale.html"
    cmd_options = {
        'orientation': 'landscape',
    }

    def get_context_data(self, **kwargs):
        cod_etp = self.kwargs.get('step', None)
        cod_anu = self.request.GET.get('cod_anu', 2015)

        etudiants = Individu.objects.using('oracle').filter(etapes__cod_etp=cod_etp,
                                            etapes__cod_anu=cod_anu, etapes__eta_iae='E',
                                            etapes__tem_iae_prm='O', ).order_by('lib_nom_pat_ind')

        context = super(ExtrationEtiquettesView, self).get_context_data(**kwargs)
        context["etudiants"] = etudiants

        return context



class ExtrationPalView(BaseAdminView):

    def get(self, request, *args, **kwargs):
        cod_etp = kwargs.get('step', '')
        step = SettingsEtape.objects.get(cod_etp=cod_etp)
        wb = xlwt.Workbook()
        ws = wb.add_sheet('etudiant')
        queryset = step.wish_set.filter(annee__cod_anu=2015, is_reins=False, state='candidature', etape_dossier__to_state='candidature_reception').order_by('individu__last_name')
        ws.write(1, 0, "code etudiant")
        ws.write(1, 1, "nom")
        ws.write(1, 2, "prenom")
        ws.write(1, 3, "code_dossier")
        ws.write(1, 4, "email")
        ws.write(1, 5, "moyenne generale")
        ws.write(1, 6, "note memoire")
        ws.write(1, 7, "note stage")
        for row, wish in enumerate(queryset):
            ws.write(row+2, 0, wish.individu.student_code)
            ws.write(row + 2, 1, wish.individu.last_name)
            ws.write(row + 2, 2, wish.individu.first_name1)
            ws.write(row + 2, 3, wish.code_dossier)
            ws.write(row + 2, 4, wish.individu.personal_email)
            try:
                ws.write(row + 2, 5, wish.notemastermodel.moyenne_general)
                ws.write(row + 2, 6, wish.notemastermodel.note_memoire)
                ws.write(row + 2, 7, wish.notemastermodel.note_stage)
            except NoteMasterModel.DoesNotExist:
                pass

        response = HttpResponse(content_type='application/vnd.ms-excel')
        date = datetime.datetime.today().strftime('%d-%m-%Y')
        response['Content-Disposition'] = 'attachment; filename=%s_%s.xls' % ('extraction', date)
        wb.save(response)
        return response

xadmin.site.register_view(r'^extraction_note_view/(?P<step>\w+)/$', ExtrationPalView, 'extraction_note')
